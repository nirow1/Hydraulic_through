import csv
import os
from datetime import datetime, timedelta

import openpyxl
from PySide6.QtCore import Signal

from PySide6.QtGui import QIcon
from PySide6.QtWidgets import QWidget, QFileDialog, QLineEdit, QTableWidgetItem

from Qt_files.Qt_python.ui_settings_view import Ui_Form


class SettingsView(QWidget):
    GENERATE_TEST_PLAN = Signal()
    SAVE_PATH = Signal(str)
    def __init__(self, logo):
        QWidget.__init__(self)
        self.ui = Ui_Form()
        self.ui.setupUi(self)

        self.xls_path: str= ""

        self.path = "./"

        self._initial_graphical_changes()
        self._bind_buttons()

    def _initial_graphical_changes(self):
        self._resize_table_widget()
        self.ui.xml_file_dir_btn.setIcon(QIcon("./App_data/dir_icon.png"))
        self.ui.save_path_dir_btn.setIcon(QIcon("./App_data/dir_icon.png"))
        self.ui.tableWidget.verticalHeader().setVisible(False)

    def _bind_buttons(self):
        self.ui.xml_file_dir_btn.clicked.connect(lambda: self._open_file_dialog(self.ui.xml_file_dir_le))
        self.ui.generate_test_plan_btn.clicked.connect(self.generate_test_plan)
        self.ui.add_row_btn.clicked.connect(self._add_row)
        self.ui.delete_row_btn.clicked.connect(self._delete_row)

        self.ui.set_saving_path_btn.clicked.connect(self._create_save_path)
        self.ui.save_path_dir_btn.clicked.connect(self._open_dir_dialog)

    def _open_file_dialog(self, lineedit: QLineEdit):
        options = QFileDialog(self).options()
        self.xls_path = QFileDialog.getOpenFileName(self, "Select Folder", "", options=options)
        lineedit.setText(self.xls_path[0])
        self._load_xls_data()

    def _open_dir_dialog(self):
        options = QFileDialog(self).options()
        folder_path = QFileDialog.getExistingDirectory(self, "Select Folder", "", options=options)
        self.ui.set_saving_path_le.setText(folder_path)

    def _load_xls_data(self):
        workbook = openpyxl.load_workbook(self.xls_path[0])
        sheet = workbook["List 1"]

        rows = sheet.max_row - 1
        self.ui.tableWidget.setRowCount(rows)

        for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column)):
            for j, cell in enumerate(row):
                cell = cell.value if cell.value is not None else ""
                self.ui.tableWidget.setItem(i, j, QTableWidgetItem(str(cell)))

        workbook.close()
        self._resize_table_widget()

    def generate_test_plan(self):
        self._save_process_dates()
        self.GENERATE_TEST_PLAN.emit()

    def _create_save_path(self):
        name = "/vysledky_" + datetime.now().strftime("%Y-%m-%d_%H_%M")
        path = self.ui.set_saving_path_le.text() + name
        os.makedirs(path)

        self.path = path
        self.SAVE_PATH.emit(path)

    def _save_process_dates(self):
        test_plan_settings = [
            [
                (item.text() if item else "")
                for item in (self.ui.tableWidget.item(row, col) for col in range(self.ui.tableWidget.columnCount()))
            ]
            for row in range(self.ui.tableWidget.rowCount())
        ]

        self._delete_data_from_files()

        with open("./App_data/Test_plan/planed_flow.csv", "a", newline="") as flow_file:
            writer = csv.writer(flow_file)
            for row in test_plan_settings:
                if row[3] is not None and row[3] != "":
                    writer.writerow([self.create_datetime(row), row[3], 0])

        self._create_cam_plans_tab("./App_data/Test_plan/cam_plans.csv", test_plan_settings)
        self._create_cam_plans_tab(self.path, test_plan_settings)

    def _create_cam_plans_tab(self, path: str, plan):
        with open(path, "a", newline="") as cam_file:
            writer = csv.writer(cam_file)
            for row in plan:
                tasks = [
                    ("foto 1", row[6] != "", row[6]),
                    ("foto 2", row[7] != "", row[7]),
                    ("video 1", row[8] != "", row[8]),
                    ("video 2", row[9] != "", row[9]),
                    ("orthophoto", row[5] == "1", 0),
                    ("photogrm", row[4] == "1", 0),
                ]
                for label, condition, value in tasks:
                    if condition:
                        writer.writerow([self.create_datetime(row), value, label, 0])


    def _delete_data_from_files(self):
        file_names = ["./App_data/Test_plan/planed_flow.csv",
                      "./App_data/Test_plan/cam_plans.csv"]
        for name in file_names:
            open(name, mode='w', newline='').close()

    def _resize_table_widget(self):
        for i in range(6):
            self.ui.tableWidget.setColumnWidth(i, 125)

    def _add_row(self):
        row_position = self.ui.tableWidget.rowCount()
        self.ui.tableWidget.insertRow(row_position)

    def _delete_row(self):
        row_position = self.ui.tableWidget.rowCount()
        if row_position > 0:
            self.ui.tableWidget.removeRow(row_position - 1)

    def change_button_status(self, status: bool):
        self.ui.generate_test_plan_btn.setEnabled(status)

    def create_datetime(self, data):
        try:
            days_to_add, hours, minutes = data[:3]
            result_datetime = datetime.now() + timedelta(days=int(days_to_add),hours=int(hours),minutes=int(minutes))
            return result_datetime.strftime("%d.%m.%Y %H:%M:%S")
        except Exception as e:
            return 0

